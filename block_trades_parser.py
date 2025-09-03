import re
import pdfplumber
import pandas as pd
import datetime as dt
from pathlib import Path
from tkinter import Tk, filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ------------------ PDF parsing helpers ------------------
GREEK_MONTHS = {
    "Ιανουαρίου":1,"Φεβρουαρίου":2,"Μαρτίου":3,"Απριλίου":4,"Μαΐου":5,"Μαίου":5,
    "Ιουνίου":6,"Ιουλίου":7,"Αυγούστου":8,"Σεπτεμβρίου":9,"Οκτωβρίου":10,"Νοεμβρίου":11,"Δεκεμβρίου":12
}
DOW = r"(Δευτέρα|Τρίτη|Τετάρτη|Πέμπτη|Παρασκευή|Σάββατο|Κυριακή)"
DATE_RE = re.compile(DOW + r",?\s+(\d{1,2})\s+([A-ΩA-Za-zΪΫά-ώϊϋΐΰΏΉΈΆΌΊΎΪΫ]+),\s+(\d{4})")

BLOCK_KEYWORDS = [
    "Στοιχεία Συναλλαγών Πακέτων",
    "Πίνακας Προσυμφωνημένων Συναλλαγών",
    "Χρεόγραφα Όγκος πακέτου Τιμή πακέτου Αξία πακέτου Ώρα έγκρισης",
]
TIME_RE = re.compile(r'(\d{2}:\d{2}:\d{2})\s+(\d+)$')

_DEACCENT_TABLE = str.maketrans("ΪΫάέίόύήώϊϋΐΰΆΈΊΎΌΉΏ", "ΙΥαειουηωιυιυΑΕΙΥΟΗΩ")

def norm_name(s: str) -> str:
    """Normalize company strings so names from PDF and Excel match."""
    if not isinstance(s, str):
        return s
    s = s.translate(_DEACCENT_TABLE)
    s = re.sub(r"\s+", " ", s.strip())
    s = re.sub(r"\s*\(", " (", s)
    s = re.sub(r"\((ΚΟ|KO)\)", "(KO)", s, flags=re.I)
    s = re.sub(r"\((ΚΑ|KA)\)", "(KA)", s, flags=re.I)
    return s.upper()

def extract_report_date(pdf_path: str) -> dt.date | None:
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            text = p.extract_text() or ""
            m = DATE_RE.search(text)
            if m:
                day = int(m.group(2))
                month_name = m.group(3)
                year = int(m.group(4))
                month = GREEK_MONTHS.get(month_name)
                if not month:
                    key = month_name.translate(_DEACCENT_TABLE)
                    fallback = {
                        "Ιανουαριου":1,"Φεβρουαριου":2,"Μαρτιου":3,"Απριλιου":4,"Μαιου":5,"Ιουνιου":6,
                        "Ιουλιου":7,"Αυγουστου":8,"Σεπτεμβριου":9,"Οκτωβριου":10,"Νοεμβριου":11,"Δεκεμβριου":12
                    }
                    month = fallback.get(key)
                if month:
                    return dt.date(year, month, day)
    return None

def locate_block_trade_pages(pdf_path: str):
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, p in enumerate(pdf.pages):
            text = p.extract_text() or ""
            if any(k in text for k in BLOCK_KEYWORDS):
                pages.append(i)
    return pages

def parse_block_table_from_page(text: str) -> pd.DataFrame:
    lines = (text or "").splitlines()
    start = None
    for i, ln in enumerate(lines):
        if "Χρεόγραφα" in ln and "Όγκος" in ln and "Ώρα έγκρισης" in ln:
            start = i + 1
            break
    rows = []
    if start is None:
        return pd.DataFrame(columns=["Company","Volume","Price","Value","ApprovalTime","Note"])
    for ln in lines[start:]:
        ln = ln.strip()
        if not ln or "Σημειώσεις" in ln:
            break
        m = TIME_RE.search(ln)
        if not m:
            continue
        time = m.group(1); note = m.group(2)
        left = ln[:m.start()].strip()
        parts = left.split()
        if len(parts) < 4:
            continue
        value_str, price_str, volume_str = parts[-1], parts[-2], parts[-3]
        company = " ".join(parts[:-3])
        rows.append([company, volume_str, price_str, value_str, time, note])
    df = pd.DataFrame(rows, columns=["Company","Volume","Price","Value","ApprovalTime","Note"])
    if not df.empty:
        df["Volume"] = df["Volume"].str.replace(",","", regex=False).astype("int64")
        df["Price"]  = df["Price"].str.replace(",","", regex=False).astype("float64")
        df["Value"]  = df["Value"].str.replace(",","", regex=False).astype("float64")
    return df

def extract_block_trades(pdf_path: str) -> pd.DataFrame:
    report_date = extract_report_date(pdf_path)
    pages = locate_block_trade_pages(pdf_path)
    frames = []
    with pdfplumber.open(pdf_path) as pdf:
        for idx in pages:
            text = pdf.pages[idx].extract_text() or ""
            df = parse_block_table_from_page(text)
            if not df.empty:
                df.insert(0, "Date", report_date)
                frames.append(df)
    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame(columns=["Date","Company","Volume","Price","Value","ApprovalTime","Note"])

# ------------------ Excel helpers ------------------
def find_or_create_date_row(ws, target_date: dt.date) -> int:
    """Find row in column A that equals target_date (date only). If not found, append a row."""
    max_row = ws.max_row or 1
    for r in range(3, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        try:
            d = pd.to_datetime(val, dayfirst=True).date()
        except Exception:
            continue
        if d == target_date:
            return r
    # append
    r = max_row + 1
    ws.cell(row=r, column=1, value=dt.datetime.combine(target_date, dt.time()))
    return r

def read_company_header_positions(ws) -> dict:
    """
    Read company headers from row 2 at columns B, F, J, ... (every 4th col).
    Returns {normalized_name: (start_col_index, raw_header_text)}
    """
    positions = {}
    max_col = ws.max_column
    c = 2  # B
    while c <= max_col:
        header = ws.cell(row=2, column=c).value
        if header and str(header).strip():
            txt = str(header).strip()
            positions[norm_name(txt)] = (c, txt)
        c += 4
    return positions

def group_trades_for_formulas(df: pd.DataFrame) -> dict:
    """
    Build: {normalized_company: {"volumes":[...], "prices":[...]} }
    """
    out = {}
    for _, row in df.iterrows():
        comp = norm_name(str(row["Company"]))
        out.setdefault(comp, {"volumes": [], "prices": []})
        out[comp]["volumes"].append(int(row["Volume"]))
        out[comp]["prices"].append(float(row["Price"]))
    return out

def price_list_greek(prices: list[float]) -> str:
    """Return dash-joined prices with comma as decimal separator."""
    return "-".join(f"{p:.2f}".replace(".", ",") for p in prices)

def volume_formula(volumes: list[int]) -> str | int:
    """Return '=a+b+...' Excel formula for volumes; if only one, return the number."""
    if not volumes:
        return None
    if len(volumes) == 1:
        return volumes[0]
    return "=" + "+".join(str(v) for v in volumes)

def fill_row(ws, row_idx: int, header_pos: dict, trades: dict, date_obj: dt.date):
    """
    Under each company group, set:
      B/F/J...   -> Ημ/νία           (always set)
      +1 column  -> ΣΠΑΚΕΤΩΝ (volume formula if trades)
      +2 column  -> #ΠΑΚΕΤΩΝ (count if trades)
      +3 column  -> ΤΙΜΗ     (dash-joined prices if trades)
    """
    for comp_norm, (start_col, _hdr) in header_pos.items():
        ws.cell(row=row_idx, column=start_col, value=dt.datetime.combine(date_obj, dt.time()))
        data = trades.get(comp_norm)
        if not data:
            continue  # leave ΣΠΑΚΕΤΩΝ/#/ΤΙΜΗ blank
        vols = data["volumes"]
        prices = data["prices"]
        ws.cell(row=row_idx, column=start_col + 1, value=volume_formula(vols))
        ws.cell(row=row_idx, column=start_col + 2, value=len(vols))
        ws.cell(row=row_idx, column=start_col + 3, value=price_list_greek(prices))

def write_pdf_sheet(wb, sheet_name: str, df_pdf: pd.DataFrame, header_pos: dict):
    """
    Create (or replace) a sheet named `sheet_name` with the PDF rows and a Matched flag.
    Rows with matched companies (present in header_pos) are green; unmatched are red.
    """
    # Replace if exists
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Prepare normalized columns and matched flag
    df_show = df_pdf.copy()
    df_show["CompanyNorm"] = df_show["Company"].map(norm_name)
    df_show["Matched"] = df_show["CompanyNorm"].apply(lambda x: "Yes" if x in header_pos else "No")

    # Write header
    cols = ["Company","Volume","Price","Value","ApprovalTime","Note","Matched"]
    for j, c in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=c)

    # Styles
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Write rows
    for i, row in df_show.iterrows():
        vals = [row["Company"], int(row["Volume"]), float(row["Price"]), float(row["Value"]), row["ApprovalTime"], row["Note"], row["Matched"]]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i+2, column=j, value=v)
        # color by match
        rng = ws.iter_rows(min_row=i+2, max_row=i+2, min_col=1, max_col=len(cols))
        for r in rng:
            for cell in r:
                cell.fill = fill_green if row["Matched"] == "Yes" else fill_red

    # Autosize a bit
    for j in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(j)].width = 18

# ------------------ Main ------------------
def main():
    Tk().withdraw()
    pdf_path = filedialog.askopenfilename(title="Select ATHEX Daily PDF", filetypes=[("PDF files","*.pdf")])
    if not pdf_path:
        print("No PDF selected."); return
    xlsx_path = filedialog.askopenfilename(title="Select Block Trades Master Excel", filetypes=[("Excel","*.xlsx")])
    if not xlsx_path:
        print("No Excel selected."); return

    pdf_path = Path(pdf_path)
    xlsx_path = Path(xlsx_path)

    # 1) Parse PDF
    df = extract_block_trades(str(pdf_path))
    if df.empty:
        print("No block trades found in PDF."); return

    # Normalize company names for matching
    df_norm = df.copy()
    df_norm["Company"] = df_norm["Company"].map(str)
    df_norm["CompanyNorm"] = df_norm["Company"].map(norm_name)

    # Determine report date & yyyy-mm-dd / dd.mm.yyyy strings
    report_date = df_norm.iloc[0]["Date"]
    if isinstance(report_date, pd.Timestamp):
        report_date = report_date.date()
    date_label = report_date.strftime("%d.%m.%Y")  # for sheet name & file name

    # 2) Open workbook + map headers
    wb = load_workbook(xlsx_path)
    # Change "Master" if your sheet name is different
    if "Master" not in wb.sheetnames:
        raise ValueError("Sheet 'Master' not found in workbook.")
    ws = wb["Master"]

    row_idx = find_or_create_date_row(ws, report_date)
    headers_map = read_company_header_positions(ws)

    # 3) Build per-company trades (for formulas), then fill the Master row
    trades_map = {}
    for comp, g in df_norm.groupby("CompanyNorm"):
        vols = [int(v) for v in g["Volume"].tolist()]
        prices = [float(p) for p in g["Price"].tolist()]
        trades_map[comp] = {"volumes": vols, "prices": prices}

    fill_row(ws, row_idx, headers_map, trades_map, report_date)

    # 4) Create the PDF→Excel sheet with match colors
    #    Keep original (non-normalized) columns for display
    df_display = df.copy()  # Company, Volume, Price, Value, ApprovalTime, Note
    write_pdf_sheet(wb, date_label, df_display, headers_map)

    # 5) Save workbook under the requested new name
    new_name = f"Block Trades_updated as of {date_label}.xlsx"
    out_path = xlsx_path.with_name(new_name)
    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
